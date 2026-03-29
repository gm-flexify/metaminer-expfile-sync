"""Import FB campaign metrics from XLSX/CSV into fb_campaigns + fb_daily_insights.

Idempotent: same (campaign_id, data_date, country) overwrites completely via ON CONFLICT.
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session

from app.db.models import FbCampaign, FbDailyInsight

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column alias system
# ---------------------------------------------------------------------------

def _norm(raw: str) -> str:
    s = raw.strip().replace("\ufeff", "").lower()
    return re.sub(r"\s+", " ", s)


FIELD_ALIASES: Dict[str, str] = {}


def _reg(canonical: str, *labels: str) -> None:
    for lab in labels:
        FIELD_ALIASES[_norm(lab)] = canonical


_reg("campaign_id", "campaign id", "campaign_id", "id кампанії", "id кампании")
_reg("account_id", "account id", "ad account id", "account_id", "id облікового запису", "рекламний акаунт")
_reg("report_date", "date", "day", "дата", "день", "report date")
_reg("name", "campaign name", "name", "назва", "назва кампанії", "название кампании")
_reg("status", "status", "статус")
_reg(
    "spend", "spend", "amount spent", "amount spent (usd)", "amount spent (eur)",
    "amount spent (uah)", "витрати", "витрата", "витрачена сума", "сума витрат",
    "сума витрат (usd)",
)
_reg("impressions", "impressions", "покази", "imps")
_reg(
    "clicks", "clicks", "кліки", "клики", "кліки (усі)", "клики (все)", "clicks (all)",
)
_reg(
    "link_clicks", "link clicks", "inline link clicks", "inline_link_clicks",
    "link_clicks", "кліки по посиланню", "клики по ссылке",
)
_reg("reach", "reach", "охоплення")
_reg("frequency", "frequency", "частота")
_reg("cpm", "cpm")
_reg("ctr", "ctr")
_reg("cpc", "cpc")
_reg("leads", "leads", "ліди", "лиды")
_reg("purchases", "purchases", "покупки")
_reg("installs", "installs", "інстали", "app installs", "встановлення додатку", "установки приложения")
_reg("registrations", "registrations", "реєстрації")
_reg("subscriptions", "subscriptions", "підписки", "подписки")
_reg("unique_clicks", "unique clicks", "unique_clicks")
_reg("unique_ctr", "unique ctr", "unique_ctr")
_reg("country", "country", "країна", "страна")
_reg("currency", "currency", "валюта")
_reg("reporting_starts", "reporting starts", "початок звітності")
_reg("reporting_ends", "reporting ends", "завершення звітності")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_INVALID_CAMPAIGN = frozenset({"all", "multiple", "none", "unknown", "total", "default"})
_INVALID_ACCOUNT = frozenset({"all", "multiple", "none", "unknown", "total", "n/a", "na", "—", "-"})


def normalize_campaign_id(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(int(v)) if isinstance(v, float) and float(v) == int(v) else str(v).strip()
    s = re.sub(r"\s+", "", s)
    return s or None


def is_plausible_fb_campaign_id(cid: str) -> bool:
    if not cid or cid.lower() in _INVALID_CAMPAIGN:
        return False
    return cid.isdigit() and len(cid) >= 10


def parse_date_value(v: Any, default: Optional[date] = None) -> Optional[date]:
    if v is None or v == "":
        return default
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return default
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        serial = float(s)
        if 20000 <= serial <= 80000:
            return date(1899, 12, 30) + timedelta(days=int(serial))
    except (ValueError, OverflowError):
        pass
    return default


def _safe_int(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _safe_decimal(v: Any) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _map_headers(row: Sequence[Any]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for i, cell in enumerate(row):
        if cell is None:
            continue
        label = _norm(str(cell))
        if not label:
            continue
        canonical = FIELD_ALIASES.get(label)
        if canonical and canonical not in col_map:
            col_map[canonical] = i
    return col_map


def _row_to_record(row: Sequence[Any], col_map: Dict[str, int], width: int) -> Dict[str, Any]:
    padded = list(row) + [None] * max(0, width - len(row))
    return {k: padded[i] if i < len(padded) else None for k, i in col_map.items()}


def _row_has_values(row: Optional[Sequence[Any]]) -> bool:
    if row is None:
        return False
    return any(c is not None and str(c).strip() for c in row)


def _split_header_body(rows: List[Tuple[Any, ...]]) -> Tuple[List[Any], List[Tuple[Any, ...]]]:
    for i, row in enumerate(rows):
        if not _row_has_values(row):
            continue
        col_map = _map_headers(row)
        if "campaign_id" in col_map:
            body = [tuple(r) for r in rows[i + 1:] if _row_has_values(r)]
            return list(row), body
    return [], []


def _parse_csv(data: bytes) -> Tuple[List[Any], List[Tuple[Any, ...]]]:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text_data = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        text_data = data.decode("utf-8", errors="replace")
    rows = [tuple(line) for line in csv.reader(StringIO(text_data))]
    return _split_header_body(rows) if rows else ([], [])


def _parse_xlsx(data: bytes) -> Tuple[List[Any], List[Tuple[Any, ...]]]:
    wb = load_workbook(filename=BytesIO(data), read_only=False, data_only=True)
    try:
        order: List[str] = []
        if "Raw Data Report" in wb.sheetnames:
            order.append("Raw Data Report")
        for n in wb.sheetnames:
            if n not in order:
                order.append(n)
        for sn in order:
            rows = [tuple(r) for r in wb[sn].iter_rows(values_only=True)]
            header, body = _split_header_body(rows)
            if header:
                logger.info("FB import: sheet=%r, rows=%d", sn, len(body))
                return header, body
        return [], []
    finally:
        wb.close()


def parse_file(data: bytes, filename: str) -> Tuple[List[Any], List[Tuple[Any, ...]]]:
    name = filename.lower()
    if name.endswith(".csv"):
        return _parse_csv(data)
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(data)
    raise ValueError("Unsupported format: use .csv, .xlsx, or .xlsm")


# ---------------------------------------------------------------------------
# Import service
# ---------------------------------------------------------------------------

@dataclass
class FbImportSkipped:
    row_index: int
    reason: str
    campaign_id: Optional[str] = None


@dataclass
class FbImportResult:
    success: bool
    message: str
    campaigns_upserted: int = 0
    insights_upserted: int = 0
    skipped: List[FbImportSkipped] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    date_min: Optional[date] = None
    date_max: Optional[date] = None


def import_fb_report(
    db: Session,
    file_bytes: bytes,
    filename: str,
    default_date: Optional[date] = None,
    dry_run: bool = False,
) -> FbImportResult:
    errors: List[str] = []
    skipped: List[FbImportSkipped] = []

    try:
        header, data_rows = parse_file(file_bytes, filename)
    except Exception as e:
        return FbImportResult(success=False, message=str(e), errors=[str(e)])

    if not header:
        return FbImportResult(success=False, message="Empty file or missing header", errors=["No header"])

    col_map = _map_headers(header)
    if "campaign_id" not in col_map:
        return FbImportResult(success=False, message="Missing campaign_id column", errors=["No campaign_id"])

    max_col = max(col_map.values()) + 1
    campaign_rows: Dict[str, Dict[str, Any]] = {}
    insight_rows: List[Dict[str, Any]] = []

    for idx, row in enumerate(data_rows, start=2):
        rec = _row_to_record(row, col_map, max_col)
        cid = normalize_campaign_id(rec.get("campaign_id"))
        if not cid or not is_plausible_fb_campaign_id(cid):
            skipped.append(FbImportSkipped(idx, "invalid or missing campaign_id", campaign_id=cid))
            continue

        report_date = parse_date_value(rec.get("report_date"), default_date)
        if not report_date:
            skipped.append(FbImportSkipped(idx, "missing date", campaign_id=cid))
            continue

        # Collect campaign info (last seen wins)
        name_val = rec.get("name")
        name_s = str(name_val).strip() if name_val and str(name_val).strip() else None
        acc_raw = rec.get("account_id")
        acc_s = str(acc_raw).strip() if acc_raw and str(acc_raw).strip().lower() not in _INVALID_ACCOUNT else None

        if cid not in campaign_rows:
            campaign_rows[cid] = {"campaign_id": cid, "name": name_s, "account_id": acc_s}
        else:
            if name_s:
                campaign_rows[cid]["name"] = name_s
            if acc_s:
                campaign_rows[cid]["account_id"] = acc_s

        ctry = rec.get("country")
        country_val = None
        if ctry is not None and str(ctry).strip():
            c_norm = str(ctry).strip().upper()[:10]
            if c_norm != "ALL":
                country_val = c_norm

        insight_rows.append({
            "campaign_id": cid,
            "data_date": report_date,
            "country": country_val,
            "spend": _safe_decimal(rec.get("spend")),
            "impressions": _safe_int(rec.get("impressions")),
            "reach": _safe_int(rec.get("reach")),
            "clicks": _safe_int(rec.get("clicks")),
            "link_clicks": _safe_int(rec.get("link_clicks")) or _safe_int(rec.get("clicks")),
            "frequency": _safe_decimal(rec.get("frequency")),
            "cpm": _safe_decimal(rec.get("cpm")),
            "cpc": _safe_decimal(rec.get("cpc")),
            "ctr": _safe_decimal(rec.get("ctr")),
            "leads": _safe_int(rec.get("leads")),
            "installs": _safe_int(rec.get("installs")),
            "purchases": _safe_int(rec.get("purchases")),
            "registrations": _safe_int(rec.get("registrations")),
            "subscriptions": _safe_int(rec.get("subscriptions")),
            "unique_clicks": _safe_int(rec.get("unique_clicks")),
            "unique_ctr": _safe_decimal(rec.get("unique_ctr")),
        })

    all_dates = [r["data_date"] for r in insight_rows if r.get("data_date")]
    date_min = min(all_dates) if all_dates else None
    date_max = max(all_dates) if all_dates else None

    if dry_run:
        return FbImportResult(
            success=True,
            message="dry_run: no writes",
            campaigns_upserted=len(campaign_rows),
            insights_upserted=len(insight_rows),
            skipped=skipped,
            date_min=date_min,
            date_max=date_max,
        )

    try:
        # Upsert campaigns
        now = datetime.utcnow()
        for c in campaign_rows.values():
            stmt = insert(FbCampaign).values(
                campaign_id=c["campaign_id"],
                name=c.get("name"),
                account_id=c.get("account_id"),
                created_at=now,
                updated_at=now,
            ).on_conflict_do_update(
                index_elements=["campaign_id"],
                set_={
                    "name": insert(FbCampaign).excluded.name,
                    "account_id": insert(FbCampaign).excluded.account_id,
                    "updated_at": now,
                },
            )
            db.execute(stmt)

        # Upsert insights
        for row in insight_rows:
            stmt = insert(FbDailyInsight).values(
                campaign_id=row["campaign_id"],
                data_date=row["data_date"],
                country=row["country"],
                spend=row["spend"],
                impressions=row["impressions"],
                reach=row["reach"],
                clicks=row["clicks"],
                link_clicks=row["link_clicks"],
                frequency=row["frequency"],
                cpm=row["cpm"],
                cpc=row["cpc"],
                ctr=row["ctr"],
                leads=row["leads"],
                installs=row["installs"],
                purchases=row["purchases"],
                registrations=row["registrations"],
                subscriptions=row["subscriptions"],
                unique_clicks=row["unique_clicks"],
                unique_ctr=row["unique_ctr"],
                created_at=now,
                updated_at=now,
            ).on_conflict_do_update(
                constraint="uq_fb_insight",
                set_={
                    "spend": insert(FbDailyInsight).excluded.spend,
                    "impressions": insert(FbDailyInsight).excluded.impressions,
                    "reach": insert(FbDailyInsight).excluded.reach,
                    "clicks": insert(FbDailyInsight).excluded.clicks,
                    "link_clicks": insert(FbDailyInsight).excluded.link_clicks,
                    "frequency": insert(FbDailyInsight).excluded.frequency,
                    "cpm": insert(FbDailyInsight).excluded.cpm,
                    "cpc": insert(FbDailyInsight).excluded.cpc,
                    "ctr": insert(FbDailyInsight).excluded.ctr,
                    "leads": insert(FbDailyInsight).excluded.leads,
                    "installs": insert(FbDailyInsight).excluded.installs,
                    "purchases": insert(FbDailyInsight).excluded.purchases,
                    "registrations": insert(FbDailyInsight).excluded.registrations,
                    "subscriptions": insert(FbDailyInsight).excluded.subscriptions,
                    "unique_clicks": insert(FbDailyInsight).excluded.unique_clicks,
                    "unique_ctr": insert(FbDailyInsight).excluded.unique_ctr,
                    "updated_at": now,
                },
            )
            db.execute(stmt)

        db.commit()
    except Exception as e:
        db.rollback()
        logger.exception("FB import failed")
        return FbImportResult(success=False, message=f"DB error: {e}", errors=[str(e)])

    return FbImportResult(
        success=True,
        message="OK",
        campaigns_upserted=len(campaign_rows),
        insights_upserted=len(insight_rows),
        skipped=skipped,
        date_min=date_min,
        date_max=date_max,
    )
